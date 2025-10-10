import streamlit as st
import openpyxl
from datetime import datetime
from collections import defaultdict
import pandas as pd
import io
import pytz

# --- Funzioni di Elaborazione Dati ---

def get_current_time_rome():
    rome_tz = pytz.timezone('Europe/Rome')
    return datetime.now(rome_tz)

def get_payment_conditions(conditions_file):
    """Legge le condizioni di pagamento da condizioni.xlsx."""
    payment_conditions = {}
    try:
        workbook = openpyxl.load_workbook(conditions_file, data_only=True)
        sheet = workbook["Sheet1"]
        last_supplier_name = None
        for row in sheet.iter_rows(values_only=True):
            supplier_code = row[0]
            supplier_info = row[1]
            payment_term = row[3]

            if supplier_code and supplier_info:
                last_supplier_name = supplier_info.strip()
                if last_supplier_name and payment_term:
                    payment_conditions[last_supplier_name] = payment_term.strip()

    except Exception as e:
        st.warning(f"Errore durante la lettura di condizioni.xlsx: {e}. La colonna 'Condizioni Pagamento' non sarà aggiunta.")
    return payment_conditions

def generate_forecasting_data(input_excel_file):
    # ... (codice di generate_forecasting_data invariato) ...
    try:
        workbook = openpyxl.load_workbook(input_excel_file)
        sheet = workbook["Sheet1"]
    except Exception as e:
        st.error(f"Errore durante l'apertura del file Excel: {e}")
        return None
    suppliers_data = defaultdict(lambda: {
        "name": "", "monthly_totals": defaultdict(float),
        "antecedenti_2025_total": 0.0, "yearly_total": 0.0
    })
    current_supplier_code = None
    current_supplier_name = None
    for row in sheet.iter_rows():
        col_a_value = row[0].value if len(row) > 0 else None
        col_b_value = row[1].value if len(row) > 1 else None
        col_d_value = row[3].value if len(row) > 3 else None
        col_m_value = row[12].value if len(row) > 12 else None
        if col_a_value == "Cod. fornitore":
            current_supplier_code = col_b_value
            current_supplier_name = col_d_value
            if current_supplier_code: suppliers_data[current_supplier_code]["name"] = current_supplier_name
        elif current_supplier_code and col_a_value and isinstance(col_a_value, (str, int, float)) and str(col_a_value).strip() not in ["Cod. fornitore", "Subtotale"] and col_d_value and col_m_value is not None:
            try:
                delivery_date = None
                if isinstance(col_d_value, datetime): delivery_date = col_d_value
                elif isinstance(col_d_value, str):
                    try: delivery_date = datetime.strptime(col_d_value, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        try: delivery_date = datetime.strptime(col_d_value, "%Y-%m-%d")
                        except ValueError:
                            try: delivery_date = datetime.strptime(col_d_value, "%d/%m/%Y")
                            except ValueError: pass
                if delivery_date and delivery_date <= datetime(2025, 12, 31):
                    amount = float(str(col_m_value).replace(",", "."))
                    if delivery_date.year == 2025: suppliers_data[current_supplier_code]['monthly_totals'][delivery_date.strftime("%m")] += amount
                    elif delivery_date.year < 2025: suppliers_data[current_supplier_code]['antecedenti_2025_total'] += amount
                    suppliers_data[current_supplier_code]['yearly_total'] += amount
            except (ValueError, TypeError): pass
    return suppliers_data

def add_conto_data(report_data, conto_file):
    """Arricchisce i dati del report con le informazioni sui conti da conto.xlsx."""
    supplier_accounts = defaultdict(list)
    try:
        wb_conto = openpyxl.load_workbook(conto_file, data_only=True)
        sheet_conto = wb_conto["Foglio1"] # Assumiamo che il foglio si chiami Foglio1
        valid_prefixes = ("50.10", "50.20", "52.10", "54.10")
        last_ragione_sociale = ""
        for row in sheet_conto.iter_rows(min_row=2, values_only=True):
            if len(row) > 5:
                current_ragione_sociale = str(row[4]).strip() if row[4] is not None else ""
                conto = str(row[5]).strip() if row[5] is not None else ""
                if current_ragione_sociale:
                    last_ragione_sociale = current_ragione_sociale
                ragione_sociale_to_use = last_ragione_sociale
                if ragione_sociale_to_use and conto.startswith(valid_prefixes):
                    supplier_accounts[ragione_sociale_to_use].append(conto)
    except Exception as e:
        st.warning(f"Errore durante la lettura di conto.xlsx: {e}. Verrà mostrato il report base senza contropartita.")
        return report_data, False
    contropartita_added = False
    for code, data in report_data.items():
        supplier_name = data["name"]
        if supplier_name in supplier_accounts:
            data["Contropartita"] = ", ".join(sorted(list(set(supplier_accounts[supplier_name]))))
            contropartita_added = True
        else:
            data["Contropartita"] = "N/A"
    return report_data, contropartita_added

st.set_page_config(page_title="Report Previsioni di Costo Economico", layout="wide")
st.title("📊 Vetronaviglio s.r.l. - Report Previsioni Costo Economico")

uploaded_file = st.file_uploader("1. Carica il file `ordfor06.xlsx`", type=["xlsx"])

# Sezione opzionale per Contropartita
with st.expander("2. Aggiungi Contropartita (Opzionale)"):
    uploaded_conto = st.file_uploader("Carica `conto.xlsx`", type=["xlsx"])

# Sezione opzionale per Condizioni di Pagamento
with st.expander("3. Aggiungi Condizioni di Pagamento (Opzionale)"):
    uploaded_condizioni = st.file_uploader("Carica `condizioni.xlsx`", type=["xlsx"])

if uploaded_file:
    st.success("File `ordfor06.xlsx` caricato.")
    
    suppliers_data = generate_forecasting_data(uploaded_file)

    contropartita_added = False
    if uploaded_conto:
        st.info("File `conto.xlsx` caricato. Aggiungo colonna 'Contropartita' al report.")
        suppliers_data, contropartita_added = add_conto_data(suppliers_data, uploaded_conto)

    payment_conditions_added = False
    if uploaded_condizioni:
        st.info("File `condizioni.xlsx` caricato. Aggiungo colonna 'Condizioni Pagamento' al report.")
        payment_conditions = get_payment_conditions(uploaded_condizioni)
        if payment_conditions:
            payment_conditions_added = True
            for code, data in suppliers_data.items():
                supplier_name = data["name"]
                data["Condizioni Pagamento"] = payment_conditions.get(supplier_name, "N/A")

    if suppliers_data:
        all_supplier_names_raw = sorted([data["name"] for data in suppliers_data.values()])
        all_supplier_names_for_multiselect = ["Tutti"] + all_supplier_names_raw
        selected_supplier_names = st.multiselect("Filtra Fornitori", options=all_supplier_names_for_multiselect, default=["Tutti"])

        if "Tutti" in selected_supplier_names: filtered_suppliers_data = suppliers_data
        elif selected_supplier_names: filtered_suppliers_data = {c: d for c, d in suppliers_data.items() if d["name"] in selected_supplier_names}
        else: filtered_suppliers_data = {}

        report_rows = []
        sorted_suppliers = sorted(filtered_suppliers_data.items(), key=lambda item: item[1]['name'])
        italian_month_names = ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]
        
        for code, data in sorted_suppliers:
            row_data = {}
            row_data["Fornitore"] = data["name"]
            row_data["Codice Fornitore"] = code
            if contropartita_added:
                row_data["Contropartita"] = data.get("Contropartita", "N/A")
            if payment_conditions_added:
                row_data["Condizioni Pagamento"] = data.get("Condizioni Pagamento", "N/A")
            row_data["Antecedenti 2025"] = data["antecedenti_2025_total"]
            for month_num in range(1, 13):
                row_data[italian_month_names[month_num - 1]] = data["monthly_totals"][f"{month_num:02d}"]
            row_data["Totale Anno"] = data["yearly_total"]
            report_rows.append(row_data)
        
        columns_order = ["Fornitore", "Codice Fornitore"]
        if contropartita_added:
            columns_order.append("Contropartita")
        if payment_conditions_added:
            columns_order.append("Condizioni Pagamento")
        columns_order.extend(["Antecedenti 2025"] + italian_month_names + ["Totale Anno"])

        df = pd.DataFrame(report_rows, columns=columns_order)

        st.dataframe(df.style.format({col: "{:,.0f} €" for col in df.columns if col not in ["Fornitore", "Codice Fornitore", "Contropartita", "Condizioni Pagamento"]}), use_container_width=True)

        output_excel_buffer = io.BytesIO()
        df.to_excel(output_excel_buffer, index=False, sheet_name='Report Previsioni')
        output_excel_buffer.seek(0)

        st.download_button(
            label="📥 Scarica Report Excel",
            data=output_excel_buffer,
            file_name="forecasting_completo.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.warning("Nessun dato generato.")

st.markdown(f"---")
st.info(f"Ultimo aggiornamento: {get_current_time_rome().strftime('%d/%m/%Y %H:%M:%S')}")
