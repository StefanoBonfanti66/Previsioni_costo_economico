import openpyxl
from datetime import datetime
from collections import defaultdict
import os
import pytz

def get_conto_data(suppliers_data, conto_path):

    """Arricchisce i dati dei fornitori con le informazioni sui conti da conto.xlsx."""

    supplier_accounts = defaultdict(list)

    try:

        wb_conto = openpyxl.load_workbook(conto_path, data_only=True)

        sheet_conto = wb_conto["Foglio1"] # Assumiamo che il foglio si chiami Foglio1

        

        # Prefissi dei conti da considerare

        valid_prefixes = ("50.10", "50.20", "52.10", "54.10")



        last_ragione_sociale = "" # Variabile per gestire le celle unite



        for row_idx, row in enumerate(sheet_conto.iter_rows(min_row=2, values_only=True)):

            if len(row) > 5:

                current_ragione_sociale = str(row[4]).strip() if row[4] is not None else ""

                conto = str(row[5]).strip() if row[5] is not None else ""



                # Gestione delle celle unite: se la ragione sociale corrente è vuota, usa l'ultima non vuota

                if current_ragione_sociale:

                    last_ragione_sociale = current_ragione_sociale

                

                ragione_sociale_to_use = last_ragione_sociale



                if ragione_sociale_to_use and conto.startswith(valid_prefixes):

                    supplier_accounts[ragione_sociale_to_use].append(conto)

    except Exception as e:

        print(f"ERRORE in get_conto_data durante la lettura di conto.xlsx: {e}")

        return suppliers_data, False # In caso di errore, ritorna i dati originali e indica che non è stato aggiunto nulla



    contropartita_added = False

    for code, data in suppliers_data.items():

        supplier_name = data["name"]

        if supplier_name in supplier_accounts:

            data["Contropartita"] = ", ".join(sorted(list(set(supplier_accounts[supplier_name])))) # Unisci conti unici e ordinati

            contropartita_added = True

        else:

            data["Contropartita"] = "N/A" # Nessun conto trovato per questo fornitore

            

    return suppliers_data, contropartita_added



def generate_forecasting_report(input_filepath, output_filepath, sheet_name="Sheet1"):

    """Genera un report di previsione e, se possibile, lo arricchisce con la contropartita."""

    try:

        workbook = openpyxl.load_workbook(input_filepath)

        sheet = workbook[sheet_name]

    except Exception as e:

        return f"Errore durante l'apertura del file di input: {e}"



    # ... (Logica di estrazione dati da ordfor06.xlsx - invariata) ...

    suppliers_data = defaultdict(lambda: {

        "name": "", "monthly_totals": defaultdict(float),

        "antecedenti_2025_total": 0.0, "yearly_total": 0.0

    })

    current_supplier_code = None

    for row in sheet.iter_rows():

        col_a_value = row[0].value if len(row) > 0 else None

        col_b_value = row[1].value if len(row) > 1 else None

        col_d_value = row[3].value if len(row) > 3 else None

        col_m_value = row[12].value if len(row) > 12 else None

        if col_a_value == "Cod. fornitore":

            current_supplier_code = col_b_value

            if current_supplier_code: suppliers_data[current_supplier_code]["name"] = row[3].value

        elif current_supplier_code and col_a_value and isinstance(col_a_value, (str, int, float)) and str(col_a_value).strip() not in ["Cod. fornitore", "Subtotale"] and col_d_value and col_m_value is not None:

            try:

                delivery_date = None

                if isinstance(col_d_value, datetime): delivery_date = col_d_value

                elif isinstance(col_d_value, str):

                    try: delivery_date = datetime.strptime(col_d_value, "%Y-%m-%d %H:%M:%S")

                    except ValueError: 

                        try: delivery_date = datetime.strptime(col_d_value, "%Y-%m-%d")

                        except ValueError: 

                            try: delivery_date = datetime.strptime(col_d_value, "%d/%m/%Y")

                            except ValueError: pass

                if delivery_date and delivery_date <= datetime(2025, 12, 31):

                    amount = float(str(col_m_value).replace(",", "."))

                    if delivery_date.year == 2025: suppliers_data[current_supplier_code]['monthly_totals'][delivery_date.strftime("%m")] += amount

                    elif delivery_date.year < 2025: suppliers_data[current_supplier_code]['antecedenti_2025_total'] += amount

                    suppliers_data[current_supplier_code]['yearly_total'] += amount

            except (ValueError, TypeError): pass



    # --- Integrazione Logica Contropartita da conto.xlsx ---
    contropartita_added = False
    conto_path = "conto.xlsx"
    if os.path.exists(conto_path):
        suppliers_data, contropartita_added = get_conto_data(suppliers_data, conto_path)

    # --- Scrittura del file Excel di output ---
    report_workbook = openpyxl.Workbook()
    report_sheet = report_workbook.active
    report_sheet.title = "Report Previsioni"

    headers = ["Fornitore", "Codice Fornitore"]
    if contropartita_added:
        headers.insert(2, "Contropartita") # Inserisce in terza posizione
    headers.extend(["Antecedenti 2025"] + [f"{month_name}" for month_name in ["Gennaio", "Febbraio", "Marzo", "Aprile", "Maggio", "Giugno", "Luglio", "Agosto", "Settembre", "Ottobre", "Novembre", "Dicembre"]] + ["Totale Anno"])
    report_sheet.append(headers)

    sorted_suppliers = sorted(suppliers_data.items(), key=lambda item: item[1]['name'])

    for code, data in sorted_suppliers:
        row_data = [data["name"], code]
        if contropartita_added:
            row_data.insert(2, data.get("Contropartita", "N/A")) # Inserisce in terza posizione
        row_data.extend([data["antecedenti_2025_total"]] + [data["monthly_totals"][f"{m:02d}"] for m in range(1, 13)] + [data["yearly_total"]])
        report_sheet.append(row_data)

    # ... (formattazione e aggiunta timestamp) ...
    currency_format = '#,##0 "€"'
    start_col = 4 if contropartita_added else 3
    for col_idx in range(start_col, len(headers) + 1):
        for row_idx in range(2, report_sheet.max_row + 1):
            cell = report_sheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)): cell.number_format = currency_format

    rome_tz = pytz.timezone('Europe/Rome')
    now_rome = datetime.now(rome_tz)
    timestamp_str = now_rome.strftime('%d/%m/%Y %H:%M:%S')
    report_sheet.append([])
    report_sheet.append([f"Aggiornato al: {timestamp_str}"])

    try:
        report_workbook.save(output_filepath)
        message = f"Report generato con successo in '{output_filepath}'"
        if contropartita_added: message += " con colonna 'Contropartita'."
        return message
    except Exception as e:
        return f"Errore durante il salvataggio del report: {e}"

if __name__ == "__main__":
    result = generate_forecasting_report("ordfor06.xlsx", "forecasting.xlsx")
    print(result)